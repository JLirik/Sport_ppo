import io
from io import BytesIO

import openpyxl
from flask import *
from models import *
from datetime import timedelta
from flask_login import LoginManager, login_user, logout_user, current_user
import pandas as pd
from flask import send_file

app = Flask(__name__)
app.config['SECRET_KEY'] = 'ppo_pumpkin'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///pumpkin.db'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)
app.config['UPLOAD_FOLDER'] = 'uploads/'
db.init_app(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

with app.app_context():
    db.create_all()


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


@app.route('/', methods=['GET'])
def home():
    if current_user.is_authenticated:
        if current_user.is_admin:
            return redirect(url_for('admin_main'))
        return redirect(url_for('user_main'))
    # create_base_db()
    return render_template('startpage.html')


@app.route('/registration', methods=['GET'])
def open_registration():
    if current_user.is_authenticated:
        return redirect(url_for('admin_main')) if current_user.is_admin else redirect(url_for('user_main'))
    return render_template('registration.html')


@app.route('/Auth/reg_check', methods=['POST'])
def registration():
    data = request.get_json()
    user = data['username']
    password = data['password']
    is_admin = int(data['admin'] == 'BigKoribskyTits')

    if User.query.filter_by(name=user).first():
        return make_response('This user already exists')

    to_db_user = User(name=user,
                      password=password,
                      is_admin=is_admin)
    db.session.add(to_db_user)
    db.session.commit()

    login_user(to_db_user)
    return make_response(f'Successfully registered;{is_admin}')


@app.route('/login', methods=['GET'])
def open_login():
    if current_user.is_authenticated:
        return redirect(url_for('admin_main')) if current_user.is_admin else redirect(url_for('user_main'))
    return render_template('login.html')


@app.route('/Auth/login_check', methods=['POST'])
def login():
    data = request.get_json()
    user = data['username']
    password = data['password']

    user = User.query.filter_by(name=user).first()
    if user and user.password == password:
        login_user(user)
        return make_response(f'Successfully logged in;{int(user.is_admin)}')
    else:
        return make_response('Wrong login or password')


@app.route('/user/main', methods=['GET'])
def user_main():
    if current_user.is_authenticated:
        if current_user.is_admin:
            return redirect(url_for('admin_main'))
    else:
        return redirect(url_for('home'))

    taken_item = Take.query.filter_by(user_id=current_user.id).all()
    user_inventory = []
    for item in taken_item:
        product = Inventory.query.filter_by(id=item.inventory_id).first()
        fixes = FixRequest.query.filter_by(user_id=current_user.id, inventory_id=item.inventory_id).first()
        status = False
        if fixes:
            status = fixes.status
        user_inventory.append((product.id, product.name, product.quality, status))
    return render_template('user-main-2.0.html', inventory=user_inventory)


@app.route('/admin/main', methods=['GET'])
def admin_main():
    if current_user.is_authenticated:
        if not current_user.is_admin:
            return redirect(url_for('user_main'))
    else:
        return redirect(url_for('home'))
    taken_item = Inventory.query.all()
    all_inventory = []
    for item in taken_item:
        all_inventory.append((item.id, item.name, item.quality))
    return render_template('admin-main-2.0.html', inventory=all_inventory)


@app.route('/logout', methods=['GET'])
def logout():
    if current_user.is_authenticated:
        logout_user()
        return redirect(url_for('open_login'))
    return redirect(url_for('home'))


@app.route('/user/request_new_item', methods=['GET'])
def request_new_item():
    if current_user.is_authenticated:
        taken_ids = [subitem.inventory_id for subitem in Take.query.all()]
        orderd_ids = [subitem.inventory_id for subitem in
                      NewRequest.query.filter(NewRequest.user_id != current_user.id).all()]
        not_taken_inventory = [[item.id, item.name, item.quality,
                                NewRequest.query.filter(
                                    NewRequest.inventory_id == item.id).first().status if NewRequest.query.filter(
                                    NewRequest.inventory_id == item.id).first() else False] for item in
                               Inventory.query.all() if item.id not in taken_ids + orderd_ids]
        return render_template('user-requests-2.0.html', inventory=not_taken_inventory)
    else:
        return redirect(url_for('home'))


@app.route('/User/send_new_item', methods=['POST'])
def send_new_item():
    data = request.get_json()
    user_id = current_user.id
    item_id = data['item_id']

    ask_item = NewRequest()
    ask_item.user_id = user_id
    ask_item.inventory_id = item_id
    ask_item.quality = Inventory.query.filter(Inventory.id == item_id).first().quality
    ask_item.status = 'На рассмотрении'
    db.session.add(ask_item)
    db.session.commit()

    return make_response('Запрос на выдачу принят')


@app.route('/User/fix_item', methods=['POST'])
def fix_item():
    item_id = request.get_json()['item_id']
    item = Inventory.query.filter(Inventory.id == item_id).first()
    req = FixRequest()
    req.user_id = current_user.id
    req.inventory_id = item_id
    req.quality = item.quality
    req.status = 'На рассмотрении'
    db.session.add(req)
    db.session.commit()
    return make_response('The Best!')


@app.route('/admin/check_requests', methods=['GET'])
def check_requests():
    if current_user.is_authenticated:
        if not current_user.is_admin:
            return redirect(url_for('user_main'))
    else:
        return redirect(url_for('home'))
    taken_item = FixRequest.query.filter(FixRequest.status == 'На рассмотрении').all()
    taken_item2 = NewRequest.query.filter(NewRequest.status == 'На рассмотрении').all()
    user_requests = []
    for ask in taken_item:
        user_requests.append((ask.id, User.query.filter(User.id == ask.user_id).first().name,
                              Inventory.query.filter(Inventory.id == ask.inventory_id).first().name, ask.quality, 0))
    for ask in taken_item2:
        user_requests.append((ask.id, User.query.filter(User.id == ask.user_id).first().name,
                              Inventory.query.filter(Inventory.id == ask.inventory_id).first().name, ask.quality, 1))
    return render_template('check_requests.html', user_requests=user_requests)


@app.route('/admin/update_request_status', methods=['POST'])
def update_request_status():
    data = request.get_json()
    request_id = data.get('id')
    action = data.get('action')
    status = data.get('status')
    if status == 0:
        request_to_fix = FixRequest.query.filter_by(id=request_id).first()
        if not request_to_fix:
            return jsonify(success=False, message="Запрос не найден"), 404

        if action == 'accept':
            request_to_fix.status = 'Принята'
            inventory = Inventory.query.filter_by(id=request_to_fix.inventory_id).first()
            if inventory:
                inventory.quality = 'Новый'
        elif action == 'reject':
            request_to_fix.status = 'Отклонена'
        db.session.commit()
        return jsonify(success=True)

    elif status == 1:
        request_to_new = NewRequest.query.filter_by(id=request_id).first()
        if not request_to_new:
            return jsonify(success=False, message="Запрос не найден"), 404
        if action == 'accept':
            request_to_new.status = 'Принята'
            take = Take(user_id=request_to_new.user_id, inventory_id=request_to_new.inventory_id)
            db.session.add(take)
        elif action == 'reject':
            request_to_new.status = 'Отклонена'
        db.session.commit()
        return jsonify(success=True)


@app.route('/admin/update_inventory', methods=['POST'])
def update_inventory():
    data = request.get_json()
    item_id = data.get('item_id')
    new_name = data.get('name')
    new_quality = data.get('quality')
    inventory = Inventory.query.filter_by(id=item_id).first()
    if inventory:
        inventory.name = new_name
        inventory.quality = new_quality
        db.session.commit()
        return jsonify(success=True, message="Данные инвентаря обновлены")
    return jsonify(success=False, message="Инвентарь не найден")


@app.route('/admin/add_item', methods=['POST'])
def add_item():
    data = request.get_json()
    item_name = data['name']
    item_price = int(data['price'])
    provider = data['provider']
    adm_requests = AdminRequest()
    adm_requests.name = item_name
    adm_requests.price = item_price
    adm_requests.provider = provider
    db.session.add(adm_requests)
    db.session.commit()
    return make_response('GOOD!')


@app.route('/admin/purchases', methods=['GET'])
def purchases():
    if current_user.is_authenticated:
        if not current_user.is_admin:
            return redirect(url_for('user_main'))
    else:
        return redirect(url_for('home'))

    taken_item = AdminRequest.query.all()
    admin_requests = []
    for item in taken_item:
        admin_requests.append((item.id, item.name, item.price, item.provider))
    return render_template('purchases.html', requests=admin_requests)


@app.route('/admin/add_purchase', methods=['GET'])
def add_purchase():
    if current_user.is_authenticated:
        if not current_user.is_admin:
            return redirect(url_for('user_main'))
    else:
        return redirect(url_for('home'))

    return render_template('add_purchase.html')


@app.route('/admin/create_report', methods=['GET'])
def create_report():
    take = Take.query.all()
    i = 2
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Имя клиента'
    ws['B1'] = 'Название инвентаря'
    ws['C1'] = 'Состояние инвентаря'
    for item in take:
        ws[f'A{i}'] = User.query.filter(User.id == item.user_id).first().name
        ws[f'B{i}'] = Inventory.query.filter(Inventory.id == item.inventory_id).first().name
        ws[f'C{i}'] = Inventory.query.filter(Inventory.id == item.inventory_id).first().quality
        i += 1
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = make_response(
        send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                  as_attachment=True, download_name='example.xlsx'))
    response.headers['Content-Disposition'] = 'attachment; filename=response.xlsx'
    return response


def create_base_db():
    to_db_user = User(name='Bob',
                      password='BlaBla',
                      is_admin=0)
    db.session.add(to_db_user)
    to_db_user = User(name='Bob1',
                      password='BlaBla',
                      is_admin=0)
    db.session.add(to_db_user)
    to_db_user = User(name='BOB_ADMIN',
                      password='LoveAlice',
                      is_admin=1)
    db.session.add(to_db_user)
    to_db_inventory = Inventory(name='Мяч 1',
                                quality='Сломанный')
    db.session.add(to_db_inventory)
    to_db_inventory = Inventory(name='Мяч 2',
                                quality='Новый')
    db.session.add(to_db_inventory)
    to_db_inventory = Inventory(name='Мяч 3',
                                quality='Новый')
    db.session.add(to_db_inventory)
    to_db_inventory = Inventory(name='Гиря 1',
                                quality='Сломанный')
    db.session.add(to_db_inventory)
    to_db_inventory = Inventory(name='Гиря 2',
                                quality='Новый')
    db.session.add(to_db_inventory)
    to_db_inventory = Inventory(name='Гиря 3',
                                quality='Новый')
    db.session.add(to_db_inventory)
    to_db_inventory = Inventory(name='Гантеля 1',
                                quality='Сломанный')
    db.session.add(to_db_inventory)
    to_db_inventory = Inventory(name='Гантеля 2',
                                quality='Новый')
    db.session.add(to_db_inventory)
    to_db_inventory = Inventory(name='Гантеля 3',
                                quality='Новый')
    db.session.add(to_db_inventory)

    to_db_inventory = AdminRequest(name='Гантеля 1',
                                   price='100', provider='Спортмастер')
    db.session.add(to_db_inventory)
    to_db_inventory = AdminRequest(name='Гантеля 2',
                                   price='10', provider='Декатлон')
    db.session.add(to_db_inventory)
    to_db_inventory = AdminRequest(name='Гантеля 3',
                                   price='30', provider='DeSport')
    db.session.add(to_db_inventory)

    to_db_take = Take(user_id=1,
                      inventory_id=1)
    db.session.add(to_db_take)
    to_db_take = Take(user_id=1,
                      inventory_id=2)
    db.session.add(to_db_take)
    to_db_take = Take(user_id=2,
                      inventory_id=3)
    db.session.add(to_db_take)
    db.session.commit()


if __name__ == '__main__':
    app.run(port=1024, host='127.0.0.1')
